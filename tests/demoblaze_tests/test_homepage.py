import pytest
import allure
import openpyxl
from pages.demoblaze.home_page import HomePage


class Util:
    @classmethod
    def getExcelData(cls):
        list = []
        path = "/Users/karthikkallu/Downloads/selenium-python-framework-master/utils/search_flights_input_data.xlsx"
        wb = openpyxl.load_workbook(path)
        # sh = wb.get_sheet_by_name("login")
        sh = wb['Arkusz1']
        rows = sh.max_row
        for r in range(2, rows + 1):
            username = sh.cell(r, 1).value
            password = sh.cell(r, 2).value
            tuple = (username, password)
            list.append(tuple)
        return list


@pytest.mark.usefixtures("setup")
class TestHomePage:
    @allure.title("Home page - smoke test")
    @allure.description("Check if home page of Demoblaze has correct title")
    @pytest.mark.parametrize("cabin_class,location_from", Util.getExcelData())
    def test_homepage_title(self, cabin_class, location_from):
        homepage = HomePage(self.driver)
        homepage.open()
        homepage.clickSignInLink()
        homepage.enterUserName(cabin_class)
        homepage.enterPassword(location_from)
        assert ("STORE" in homepage.get_page_title())
